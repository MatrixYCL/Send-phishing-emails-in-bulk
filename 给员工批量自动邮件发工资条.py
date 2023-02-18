################### 实战-给员工批量自动邮件发工资条  #######################
from openpyxl import load_workbook
import smtplib

from email.mime.text import MIMEText
from email.header import Header

# 加载excel文件
wb = load_workbook("工资条.xlsx",data_only=True)

sheet = wb.active

# 设置邮箱服务器地址以及端口
smtp_server = "smtp.qq.com"
smtp_obj = smtplib.SMTP(smtp_server, 25)
smtp_obj.login("xxxx@qq.com","xxxkebxxfc")
# smtp_obj.set_debuglevel(1)      				# 显示调试信息 (辅助)

# 循环excel
count = 0
table_col_html='<thead>'    # 表头
for row in sheet.iter_rows(min_row=1):
    count+=1
    if count==1:            # first row
        for col in row:
            table_col_html+=f"<th>{col.value}</th>"
        table_col_html+="</thead>"
        continue
    else:
        row_text="<tr>"             #开始一行
        for cell in row:
            row_text+=f"<td>{cell.value}</td>"
        row_text+="</tr>"           # 结束一行
        name=row[2]
        staff_email = row[1].value
        print(staff_email,name.value)

    mail_body_context=f"""
        <h3>{name.value},  你好：</h3>
        <p>请查收你的2023-03月的工资条，钱很多。。。</p>
        <table border="lpx solid black">
            {table_col_html}
        {row_text}
        </table>
    """
    msg_body=MIMEText(mail_body_context,'html',"utf-8")

    msg_body['From'] = Header("人事部","utf-8")        # 发送者
    msg_body['To'] = Header("大唐员工","utf-8")         # 接收者
    msg_body["Subject"] = Header("大唐建设集团2023-03工资","utf-8")     # 主题

    # 发邮件
    smtp_obj.sendmail("xxxx6@qq.com",[staff_email,],msg_body.as_string())
    print(f"成功发送工资条到{staff_email}-{name.value}...")
